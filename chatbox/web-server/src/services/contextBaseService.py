import re
import ast
import math
import csv
import gevent
import chardet
from src.constants.http_status_codes import (
    HTTP_200_OK,
    HTTP_500_INTERNAL_SERVER_ERROR,
    HTTP_400_BAD_REQUEST,
    HTTP_404_NOT_FOUND
)
from nomic import atlas, AtlasDataset
from src.models import Plug, MapPoint, User, ContextItem, MapItem
from bson import ObjectId
from io import StringIO, BytesIO
from src.routes.ws import statusUpdate
from openpyxl import load_workbook
from src.helper.reader import csv_reader, xls_reader
from src.helper.file_reader import FileReader
from src.services.visualizerService import visualize_embeddings
from io import BytesIO
import base64
from gridfs import GridFS

from mongoengine.connection import get_db
import src.database.mongodb as mongo
mongo.create_db_connection()

def handler_create_map_point(provider, user_id, plug, context_base_id, feature, index=None, fs=None):
    allowed_formats = ["csv", "json", "xlsx", "xls"]
    user = User.objects(id=user_id).first()
    if not context_base_id:
        return {"code": 400, "message": "Missing context base id"}, HTTP_400_BAD_REQUEST
    if not user:
        return {"code": 404, "message": "User not found"}, HTTP_404_NOT_FOUND
    if not plug:
        return {"code": 404, "message": "Plug not found"}, HTTP_404_NOT_FOUND
    if not plug.active:
        return {"code": 400, "message": "Plug is not active"}, HTTP_400_BAD_REQUEST
    map_point = plug.mapsPoint.filter(contextBaseId=context_base_id, provider=provider).first()
    context_item = ContextItem.objects(id=context_base_id).first()
    if map_point:
        return {"code": 400, "message": "Map already existed"}, HTTP_400_BAD_REQUEST
    if not context_item:
        return {"code": 404, "message": "Context item not found"}, HTTP_404_NOT_FOUND
    if feature not in plug.features:
        return {"code": 400, "message": "This plug doesn't have that feature"}, HTTP_400_BAD_REQUEST
    file = fs.get(ObjectId(context_base_id))
    if not file:
        return {"code": 404, "message": "File not found"}, HTTP_404_NOT_FOUND
    if file.length / math.pow(1024, 2) > 1:
        return {"code": 400, "message": "File too large"}, HTTP_400_BAD_REQUEST
    if file.fileType not in allowed_formats:
        return {"code": 400, "message": "Unsupported file type"}, HTTP_400_BAD_REQUEST
    contents = file.read()
    file.seek(0)
    detected_encoding = chardet.detect(contents)['encoding']
    if detected_encoding is not None:
        text_content = contents.decode(detected_encoding)
    else:
        text_content = contents.decode('utf-8', errors='replace')
    data = None
    if file.fileType in ['csv', 'json']:
        if file.fileType == 'csv':
            csv_file = StringIO(text_content)
            reader = csv.DictReader(csv_file)
            data = [row for row in reader]
            if not index:
                index = reader.fieldnames[0]
        elif file.fileType == 'json':
            data = json.loads(text_content)
            if data and isinstance(data, list) and isinstance(data[0], dict):
                index = index or list(data[0].keys())[0]
            # If an index is needed for JSON data, you'll need additional logic here

    elif file.fileType in ['xlsx', 'xls']:
        workbook = load_workbook(filename=BytesIO(contents), read_only=True)
        sheet = workbook.active
        headers = [cell for cell in next(sheet.iter_rows(min_row=1, max_row=1, values_only=True))]
        index = index or headers[0]
        data = []
        for row in sheet.iter_rows(min_row=2, values_only=True):
            for i, value in enumerate(row):
                cell_data = {headers[i]: value}
                data.append(cell_data)
    if data is None:
        return {"code": 400, "message": "Unable to process file contents"}, HTTP_400_BAD_REQUEST

    # try:
    #     decoded_contents = text_content.decode("utf-8")
    # except UnicodeDecodeError:
    #     return {"code": 400, "message": "Error decoding file contents"}, HTTP_400_BAD_REQUEST
    # csv_file = StringIO(decoded_contents)
    # reader = csv.DictReader(csv_file)
    # data = [row for row in reader]
    # if not index:
    #     index = reader.fieldnames[0]
    url = None
    new_file_name = file.source.split(".")[0]
    context_item.isMapCreated = False
    context_item.save()
    gevent.spawn(map_upload, provider, new_file_name, index, data, file, context_base_id, plug, context_item, url,
                 user_id)
    return {
        "code": 200,
        "data": context_item.to_json(),
        "message": "Starting to create map.",
    }, HTTP_200_OK


def map_upload(provider, new_file_name, index, data, file, context_base_id, plug, context_item, url=None, user_id=None):
    # try:
        if provider == "nomic":
            dataset = atlas.map_data(data=data,
                                     indexed_field=index,
                                     description=file.source,
                                     identifier=new_file_name,
                                     )
            map_dataset = dataset.maps[0]
            url_pattern = r"https?://\S+"
            url = re.findall(url_pattern, str(map_dataset))[0]

            context_item.mapPointUrl = url
        elif provider == "2gai":
            data_reader = FileReader().load_data(input_file=file, file_suffix=".csv",
                                metadata={
                    "file_name": new_file_name,
                    "context_id": context_base_id,
                    "plug_id": str(plug.id)
                })
            # print(data_reader[0])
            id_image = visualize_embeddings(data_reader[0], context_item=context_item, image_name=new_file_name,context_base_id=context_base_id, user_id=user_id)
            file = GridFS(get_db(), collection="MapItem").get(ObjectId(id_image))
            file_data = file.read()
            blob_object = BytesIO(file_data)
            blob_data_base64 = base64.b64encode(blob_object.getvalue()).decode('utf-8')
            context_item.mapPointUrl = blob_data_base64
            context_item.isMapImage = True


        if not url and not id_image:
            context_item.isMapCreated = None
            statusUpdate(context_item={}, is_file=True,
                         message="Failed to create map file.", room=str(user_id), event='context_item_upload')
            return
        map_point = MapPoint(
            name=new_file_name, contextBaseId=context_base_id, url=url, provider=provider, isImage=True if provider == "2gai" else False)
        plug.mapsPoint.append(map_point)
        plug.save()
        context_item.isMapCreated = True
        context_item.save()
        statusUpdate(context_item=context_item.to_json(), is_file=True,
                     message="Create map successfully.", room=str(user_id), event='context_item_upload')
    # except Exception as e:
    #     context_item.isMapCreated = None
    #     statusUpdate(context_item=context_item.to_json(), is_file=True,
    #                  message=f"Create map failed. {str(e)}", room=str(user_id), event='context_item_upload')


def delete_map_point(provider, map_point, plug, context_item):
    try:
        statusUpdate(is_file=True, context_item=context_item.to_json(),
                     message="Deleting map.", room=str(plug.userId), event='context_item_upload')
        if provider == "nomic":
            dataset = AtlasDataset(map_point.name)
            dataset.delete()
        elif provider == "2gai":
            context_item.isMapImage = False
            context_item.isMapCreated = False
            MapItem.objects(contextBaseId=context_item.id).delete()
        plug.mapsPoint.remove(map_point)
        plug.save()
        context_item.isMapCreated = False
        context_item.mapPointUrl = None
        context_item.save()
        statusUpdate(is_file=True, context_item=context_item.to_json(),
                     message="Delete map successfully.", room=str(plug.userId), event='context_item_upload')
    except Exception as e:
        print(e)
        return {"code": 500, "message": "Failed to delete map"}, HTTP_500_INTERNAL_SERVER_ERROR
